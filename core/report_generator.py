from dataclasses import dataclass
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook import Workbook
from openpyxl import utils
from core.models import TransportNoticeModel, DepozitDataModel
from core.enums import NoticeType
from core.config import APP_CONFIG, VOLUME_PRECISION, PRICE_PRECISION
from core.logger import Logger

@dataclass
class CategoryStats:
    """Aggregated, precomputed statistics for one notice category (all / intrare /
    iesire, etc). Every value is computed in Python from notice.totals and written to the
    statistics sheet as a static number — no Excel formulas."""
    notice_count: int
    # Volume totals (m³)
    v_total_lr: float       # Vol Lemn rotund
    v_total_lf: float       # Vol Lemn foc
    v_total_ch: float       # Vol Cherestele
    v_total_mat: float      # Vol total materiale (volum_total_aviz field)
    v_avg_aviz: float       # Volum mediu / aviz
    # Price totals (RON)
    p_total_m: float        # Preț total materiale
    p_total_transp: float   # Preț total transport
    p_total_expl: float     # Preț total exploatare
    p_total_aviz: float     # Preț total general
    v_total_impozit: float  # Valoare totală impozit
    # Derived averages
    avg_lr: float           # Preț mediu Lemn rotund (volume-weighted)
    avg_lf: float           # Preț mediu Lemn foc (volume-weighted)
    avg_ch: float           # Preț mediu Cherestele (volume-weighted)
    avg_mat: float          # Preț mediu materiale (materiale cost / volum)

# ------------------------------------------------------------------ #

class ReportGenerator:
    def __init__(self, logger: Logger):
        self.logger = logger

        # Columns of a notice sheet. All values are computed in Python and written
        # as static numbers — the sheet is a rendered snapshot, not a live workbook.
        self.notices_sheet_headers = [
            "#", "Cod unic", "Data Emitere", "Tip aviz", "Proveniență", "Mijloc transport",
            "V total materiale (m³)", "V Lemn rotund (m³)", "V Lemn foc (m³)", "V Cherestele (m³)",
            "Preț total materiale (RON)", "Preț transport (RON)", "Preț exploatare (RON)",
            "Preț total aviz (RON)", "Valoare impozit (0.5%)"
        ]

        # Columns of a notice sheet that are prices (1 based, Excel perspective)
        self._PRICE_COLS = frozenset({11, 12, 13, 14, 15})
        # Columns of a notice sheet that are volumes (1 based, Excel perspective)
        self._VOL_COLS = frozenset({7, 8, 9, 10})
        # Columns in the notice sheet that have totals calculated for them (1 based, Excel perspective)
        self._SUM_COLS = frozenset({7, 8, 9, 10, 11, 12, 13, 14, 15})

        # Excel number formats — price uses PRICE_PRECISION decimals; volume shows up
        # to VOLUME_PRECISION decimals and renders an exact zero as "0" (the ";;0"
        # section) instead of blank.
        # Price uses an explicit positive;negative;zero format so an exact zero shows
        # as a bare "0" (the final ";0" section) rather than "0.00"; positive and
        # negative keep PRICE_PRECISION decimals.
        _price_dec = '0' * PRICE_PRECISION
        self._FMT_PRICE = f'#,##0.{_price_dec};-#,##0.{_price_dec};0'
        self._FMT_VOLUME = '#,##0.' + '#' * VOLUME_PRECISION + ';;0'

        # Fonts
        self.calibri_12 = Font(name = "Calibri", size = 12)
        self.calibri_12_bold = Font(name = "Calibri", size = 12, bold = True)
        self.calibri_12_bold_white = Font(name = "Calibri", size = 12, bold = True, color = "FFFFFF")
        self.stats_font_header = Font(name = "Calibri", size = 14, bold = True, color = "FFFFFF")
        self.stats_section_font = Font(name = "Calibri", size = 12, bold = True, italic = True)
        self.no_notices_font = Font(name = "Calibri", size = 12, italic = True, color = "808080")

        # Alignments
        self.center_align = Alignment(horizontal = "center", vertical = "center")
        self.header_align = Alignment(horizontal = "center", vertical = "center", wrap_text = True)

        # Fills
        self.light_blue_fill = PatternFill(start_color = "E5F3FF", end_color = "E5F3FF", fill_type = "solid")
        self.dark_blue_fill = PatternFill(start_color = "0067C0", end_color = "0067C0", fill_type = "solid")
        self.stats_title_fill = PatternFill(start_color = "003D82", end_color = "003D82", fill_type = "solid")

        # Borders
        self.thin_border = Border(
            left = Side(style = "thin"), right = Side(style = "thin"),
            top = Side(style = "thin"), bottom = Side(style = "thin")
        )

        # Notice type sets — reused across sheet filters and price logic
        self.INTRARE_T = frozenset({
            NoticeType.INTRARE_DIN_PARTIDA_PROPRIE,
            NoticeType.INTRARE_DIN_SURSA_EXTERNA
        })
        self.IESIRE_T = frozenset({
            NoticeType.IESIRE_DIN_DEPOZIT_PRINCIPAL,
            NoticeType.IESIRE_DIN_DEPOZIT_LR
        })

        # Sheet name constants
        self.SH_DD = "Date depozite"
        self.SH_ALL = "Toate avizele"
        self.SH_IN = "Avize intrare"
        self.SH_OUT = "Avize iesire"
        self.SH_PIN = "Avize prestari intrare"
        self.SH_POUT = "Avize prestari iesire"
        self.SH_EIN  = "Avize exploatare intrare"
        self.SH_EOUT = "Avize exploatare iesire"

    # ---------------------------------------------------------------------- #

    def _populate_notice_totals(self, notice: TransportNoticeModel, deposit_data: DepozitDataModel) -> None:
        """Populate totals for each notice"""
        prices = deposit_data.price_data
        notice_type = notice.type

        # Determine prices, if input -> pret intrare, if output = pret iesire
        is_intrare: bool = notice_type in self.INTRARE_T
        if is_intrare:
            notice.totals.pret_mc_lemn_rotund = prices.p_intrare_lr_mc
            notice.totals.pret_mc_lemn_foc = prices.p_intrare_lf_mc
            notice.totals.pret_mc_cherestele = prices.p_intrare_ch_mc
        else:
            notice.totals.pret_mc_lemn_rotund = prices.p_iesire_lr_mc
            notice.totals.pret_mc_lemn_foc = prices.p_iesire_lf_mc
            notice.totals.pret_mc_cherestele = prices.p_iesire_ch_mc

        # Calculate total prices for each good type
        notice.totals.pret_total_lemn_rotund = (
            notice.totals.pret_mc_lemn_rotund * notice.totals.volum_total_lemn_rotund
        )
        notice.totals.pret_total_lemn_foc = (
            notice.totals.pret_mc_lemn_foc * notice.totals.volum_total_lemn_foc
        )
        notice.totals.pret_total_cherestele = (
            notice.totals.pret_mc_cherestele * notice.totals.volum_total_cherestele
        )

        # Calculate total good price
        notice.totals.pret_total_materiale = (
            notice.totals.pret_total_lemn_rotund +
            notice.totals.pret_total_lemn_foc +
            notice.totals.pret_total_cherestele
        )

        # Store and calculate pret transport
        notice.totals.pret_mc_transport = prices.p_transport_mc
        notice.totals.pret_total_transport = (
            notice.volum_total_aviz * notice.totals.pret_mc_transport
        )

        total_price: float = notice.totals.pret_total_materiale + notice.totals.pret_total_transport

        # Calculate pret exploatare if that's the case
        if notice_type in [NoticeType.INTRARE_DIN_PARTIDA_PROPRIE,
                           NoticeType.IESIRE_DIN_DEPOZIT_LR]:

            notice.totals.pret_mc_exploatare = prices.p_expl_mc
            notice.totals.pret_total_exploatare = (
                notice.volum_total_aviz * notice.totals.pret_mc_exploatare
            )

            total_price += notice.totals.pret_total_exploatare

        # Populate total notice price
        notice.totals.pret_total_aviz = total_price

        # Calculate imposit value if that's the case
        if notice.type == NoticeType.INTRARE_DIN_PARTIDA_PROPRIE:
            notice.totals.valoare_impozit_punere_in_piata = (
                (notice.totals.pret_total_materiale + notice.totals.pret_total_exploatare) * APP_CONFIG.TAX_RATE_IMPOZIT
            )
        elif notice.type == NoticeType.IESIRE_DIN_DEPOZIT_LR:
            notice.totals.valoare_impozit_punere_in_piata = (
                notice.totals.pret_total_materiale * APP_CONFIG.TAX_RATE_IMPOZIT
            )

    # ---------------------------------------------------------------------- #

    def _get_col_nr_format(self, col_idx: int) -> str | None:
        """Determines the format for an Excel cell"""
        if col_idx in self._PRICE_COLS:
            return self._FMT_PRICE
        if col_idx in self._VOL_COLS:
            return self._FMT_VOLUME
        return None

    # ---------------------------------------------------------------------- #

    def _apply_styles_to_cell(
            self, cell: Cell, font: Font, fill: PatternFill = None, 
            alignment: Alignment = None, nr_format: str = None
        ) -> None:
        """Applies styling to a cell, uses default center align if no alignment is passed
        and does not set fill if it is not passed"""
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = alignment if alignment else self.center_align
        cell.border = self.thin_border
        if nr_format:
            cell.number_format = nr_format


    # ---------------------------------------------------------------------- #

    def _apply_styles_to_sheet_header(
            self, ws: Worksheet, num_columns: int, alignment: Alignment = None
        ) -> None:
        """Applies styling to header columns"""
        align = alignment or self.center_align
        # Make sure to get to num_columns + 1, range excludes last value
        for col_idx in range(1, num_columns + 1):
            cell = ws.cell(row = 1, column = col_idx)
            self._apply_styles_to_cell(
                cell, font = self.calibri_12_bold_white,
                fill = self.dark_blue_fill, alignment = align
            )

    # ---------------------------------------------------------------------- #

    def _auto_adjust_column_widths(self, ws: Worksheet):
        """Adjust Excel column widths automatically for each column by determining the largest width"""
        CHAR_WIDTH = 1.35
        PADDING = 2.0
        MIN_WIDTH = 10
        MAX_WIDTH = 80

        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            header_text = str(col_cells[0].value or "")
            header_width = max(len(ln) for ln in header_text.split("\n")) * CHAR_WIDTH + PADDING

            content_width = 0.0
            for cell in col_cells[1:]:
                if cell.value is None or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                content_width = max(content_width, max(len(ln) for ln in str(cell.value).split("\n")) * CHAR_WIDTH + PADDING)

            ws.column_dimensions[col_letter].width = min(max(header_width, content_width, MIN_WIDTH), MAX_WIDTH)

    # ---------------------------------------------------------------------- #

    def _add_excel_table(self, ws: Worksheet, table_name: str, num_rows: int, num_columns: int) -> None:
        """Add Excel table. All sheets are Excel tables for the filtering"""
        if num_rows < 1:
            return
        last_col = utils.get_column_letter(num_columns)
        tab = Table(displayName = table_name, ref = f"A1:{last_col}{num_rows}")
        tab.tableStyleInfo = TableStyleInfo(
            name = "TableStyleMedium9", showFirstColumn = False,
            showLastColumn = False, showRowStripes = True, showColumnStripes = False)
        ws.add_table(tab)

    # ---------------------------------------------------------------------- #

    def _build_notice_row_data(self, row_idx: int, notice: TransportNoticeModel) -> list:
        """Build the static value row for a notice sheet. Every numeric column is a
        precomputed value from notice.totals — no Excel formulas."""
        t = notice.totals
        return [
            row_idx - 1,
            notice.cod_unic,
            notice.data_ora_emitere.strftime(APP_CONFIG.TIME_FORMAT_SHORT),
            notice.type,
            notice.provenienta,
            notice.cap_tractor,
            notice.volum_total_aviz,
            t.volum_total_lemn_rotund,
            t.volum_total_lemn_foc,
            t.volum_total_cherestele,
            t.pret_total_materiale,
            t.pret_total_transport,
            t.pret_total_exploatare,
            t.pret_total_aviz,
            t.valoare_impozit_punere_in_piata
        ]

    # ---------------------------------------------------------------------- #

    def _write_notices_to_sheet(
            self, ws: Worksheet, notices: list[TransportNoticeModel], table_name: str
        ) -> int:
        """Write data rows as static values + a TOTAL row that SUMs each numeric
        column. Returns the notice count."""

        # Write each notice as a row, starting from row 2 (row 1 is the header)
        for row_idx, notice in enumerate(notices, start = 2):
            row_data = self._build_notice_row_data(row_idx, notice)

            for col_idx, value in enumerate(row_data, start = 1):
                cell = ws.cell(row = row_idx, column = col_idx, value = value)
                self._apply_styles_to_cell(
                    cell, font = self.calibri_12, 
                    nr_format = self._get_col_nr_format(col_idx)
                )

        # If no notices, show a placeholder message and return early
        if not notices:
            cell = ws.cell(row = 2, column = 1, value = "Nici un aviz de afișat")
            cell.font = self.no_notices_font
            cell.alignment = self.center_align
            self._auto_adjust_column_widths(ws)
            return 0

        # Row positions for the TOTAL row (right after the last data row)
        total_row = len(notices) + 2
        first_data_row = 2
        last_data_row = total_row - 1

        # Write "TOTAL" label, then a SUBTOTAL formula per numeric column. Function
        # 109 = "sum, excluding rows hidden by a filter", so when the user filters the
        # table (e.g. only iesire-din-depozit-principal in "Toate avizele"), the TOTAL
        # row updates to match the visible rows. With no filter it equals a plain SUM.
        ws.cell(row = total_row, column = 1, value = "TOTAL")
        for col in self._SUM_COLS:
            col_letter = utils.get_column_letter(col)
            ws.cell(
                row = total_row, column = col,
                value = f"=SUBTOTAL(109, {col_letter}{first_data_row}:{col_letter}{last_data_row})"
            )

        # Apply bold styling + blue background to the entire total row, with number formats
        for col_idx in range(1, len(self.notices_sheet_headers) + 1):
            cell = ws.cell(row = total_row, column = col_idx)
            self._apply_styles_to_cell(
                cell, font = self.calibri_12_bold, fill = self.light_blue_fill,
                nr_format = self._get_col_nr_format(col_idx)
            )

        # Wrap header + data rows in an Excel Table for filtering/sorting. The
        # TOTAL row is deliberately OUTSIDE the table ref: inside it, a filter
        # would treat it as a data row and hide it instead of re-summing.
        self._add_excel_table(ws, table_name, len(notices) + 1, len(self.notices_sheet_headers))

        # Resize columns to fit content
        self._auto_adjust_column_widths(ws)
        return len(notices)

    # ---------------------------------------------------------------------- #

    def _create_deposit_data_sheet(self, wb: Workbook, deposit_data: list[DepozitDataModel]) -> None:
        """Creates deposit data sheet writing header and data"""
        ws = wb.create_sheet("Date depozite", 0)
        headers = [
            "Nume\ndepozit", "Tip\ndepozit", "Sursă\ndepozit",
            "Preț intrare\nLemn rotund / m³", "Preț intrare\nLemn foc / m³",
            "Preț intrare\nCherestele / m³",
            "Preț ieșire\nLemn rotund / m³", "Preț ieșire\nLemn foc / m³",
            "Preț ieșire\nCherestele / m³",
            "Preț\ntransport / m³", "Preț\nexploatare / m³"
        ]

        for col_idx, header in enumerate(headers, 1):
            ws.cell(row = 1, column = col_idx, value = header)

        self._apply_styles_to_sheet_header(ws, len(headers), alignment = self.header_align)

        for row_idx, deposit in enumerate(deposit_data, start = 2):
            prices = deposit.price_data
            row_data = [
                deposit.nume_depozit, deposit.tip_depozit, deposit.sursa_depozit,
                prices.p_intrare_lr_mc, prices.p_intrare_lf_mc, prices.p_intrare_ch_mc,
                prices.p_iesire_lr_mc, prices.p_iesire_lf_mc, prices.p_iesire_ch_mc,
                prices.p_transport_mc, prices.p_expl_mc
            ]

            for col_idx, value in enumerate(row_data, 1):
                self._apply_styles_to_cell(
                    ws.cell(row = row_idx, column = col_idx, value = value), font = self.calibri_12
                )
        self._add_excel_table(ws, "TabelDateDepozite", len(deposit_data) + 1, len(headers))
        self._auto_adjust_column_widths(ws)

    # ---------------------------------------------------------------------- #

    def _get_deposit_for_notice(
            self, notice: TransportNoticeModel, deposit_data: list[DepozitDataModel]
        ) -> DepozitDataModel:
        """Find deposit data for a notice by matching provenienta
        to deposit name and return the whole object"""
        for deposit in deposit_data:
            if deposit.nume_depozit == notice.provenienta:
                return deposit
        # Must be found, validated by main controller before generating report starting

    # ---------------------------------------------------------------------- #

    def _make_notice_sheet(
            self, wb: Workbook, sheet_name: str, notices: list[TransportNoticeModel], table_name: str
        ) -> int:
        """Create a sheet, write the header + styling, then delegate to
        _write_notices_to_sheet for the static data + total rows. Returns notice count."""

        ws = wb.create_sheet(sheet_name)
        for col_idx, header in enumerate(self.notices_sheet_headers, start = 1):
            ws.cell(row = 1, column = col_idx, value = header)

        self._apply_styles_to_sheet_header(ws, len(self.notices_sheet_headers))

        return self._write_notices_to_sheet(ws, notices, table_name)

    # ---------------------------------------------------------------------- #

    def _create_transport_sheet(
            self, wb: Workbook, notices: list[TransportNoticeModel],
            sheet_name: str, transport_filter: str = None
        ) -> None:
        """Creates transport sheets in the workbook, one for each unique car number encountered in the notices"""
        if transport_filter:
            filtered = [n for n in notices
                        if n.cap_tractor.strip() == transport_filter]
        else:
            filtered = [n for n in notices
                        if "LIR" in n.cap_tractor.upper()]

        safe = "Tabel" + sheet_name.replace(" ", "").replace("-", "").replace("_", "")
        self._make_notice_sheet(wb, sheet_name, filtered, safe)

    # ---------------------------------------------------------------------- #

    def _aggregate_category_statistics(self, notices: list[TransportNoticeModel]) -> CategoryStats:
        """Aggregate the precomputed Python totals for one notice category into the
        values the statistics sheet displays.

        Single pass over notices to accumulate volume/price sums and
        weighted-average components (sum(vol * unit_price) / sum(vol)).
        """
        # Number of notices in this category (drives "Număr avize" + average/aviz).
        notice_count = len(notices)

        # Volume totals (m³): total aviz volume + one per good type.
        v_total_mat = v_total_lr = v_total_lf = v_total_ch = 0.0
        # Price totals (RON): materiale, transport, exploatare, grand total, impozit.
        p_total_m = p_total_transp = p_total_expl = p_total_aviz = v_impozit = 0.0

        # Volume-weighted average unit price, accumulated as cost/volume pairs:
        # avg = sum(volume * unit_price) / sum(volume). Keeping them separate
        # lets us weight each notice's price by its volume instead of averaging the
        # per-notice prices (which would ignore how much volume each notice carries).
        lr_cost_sum = lr_volume_sum = 0.0   # Lemn rotund
        lf_cost_sum = lf_volume_sum = 0.0   # Lemn foc
        ch_cost_sum = ch_volume_sum = 0.0   # Cherestele
        mat_cost_sum = mat_volume_sum = 0.0 # Materiale (combined)

        for notice in notices:
            totals = notice.totals
            # Per-notice volumes pulled out once for readability.
            volum_aviz = notice.volum_total_aviz
            volum_lr = totals.volum_total_lemn_rotund
            volum_lf = totals.volum_total_lemn_foc
            volum_ch = totals.volum_total_cherestele

            # Accumulate volume totals.
            v_total_mat += volum_aviz
            v_total_lr += volum_lr
            v_total_lf += volum_lf
            v_total_ch += volum_ch

            # Accumulate price totals (all precomputed in _populate_notice_totals).
            p_total_m += totals.pret_total_materiale
            p_total_transp += totals.pret_total_transport
            p_total_expl += totals.pret_total_exploatare
            p_total_aviz += totals.pret_total_aviz
            v_impozit += totals.valoare_impozit_punere_in_piata

            # Weighted-average components. Only notices that actually carry a good
            # type contribute to its average (guard > 0), so a notice with no Lemn
            # foc never drags that average toward zero. For materiale the weight is
            # the whole aviz volume and the cost is the materiale cost directly.
            if volum_lr > 0:
                lr_cost_sum += volum_lr * totals.pret_mc_lemn_rotund
                lr_volume_sum += volum_lr
            if volum_lf > 0:
                lf_cost_sum += volum_lf * totals.pret_mc_lemn_foc
                lf_volume_sum += volum_lf
            if volum_ch > 0:
                ch_cost_sum += volum_ch * totals.pret_mc_cherestele
                ch_volume_sum += volum_ch
            if volum_aviz > 0:
                mat_cost_sum += totals.pret_total_materiale
                mat_volume_sum += volum_aviz

        # ---------------------------------------------------------------------- #

        # Finalize a weighted average; guards against div-by-zero when a category
        # has no volume of that good type (returns 0.0 instead of erroring).
        def weighted_average(cost_sum: float, volume_sum: float) -> float:
            return round(cost_sum / volume_sum, PRICE_PRECISION) if volume_sum > 0 else 0.0

        # ---------------------------------------------------------------------- #

        # Round only at the end: volumes to VOLUME_PRECISION (6 dp, the source
        # precision) and money to PRICE_PRECISION (2 dp). v_avg_aviz is total aviz
        # volume / notice count (guarded for the empty-category case).
        return CategoryStats(
            notice_count = notice_count,
            v_total_mat = round(v_total_mat, VOLUME_PRECISION),
            v_total_lr = round(v_total_lr, VOLUME_PRECISION),
            v_total_lf = round(v_total_lf, VOLUME_PRECISION),
            v_total_ch = round(v_total_ch, VOLUME_PRECISION),
            p_total_m = round(p_total_m, PRICE_PRECISION),
            p_total_transp = round(p_total_transp, PRICE_PRECISION),
            p_total_expl = round(p_total_expl, PRICE_PRECISION),
            p_total_aviz = round(p_total_aviz, PRICE_PRECISION),
            v_total_impozit = round(v_impozit, PRICE_PRECISION),
            v_avg_aviz = round(v_total_mat / notice_count, PRICE_PRECISION) if notice_count else 0.0,
            avg_lr = weighted_average(lr_cost_sum, lr_volume_sum),
            avg_lf = weighted_average(lf_cost_sum, lf_volume_sum),
            avg_ch = weighted_average(ch_cost_sum, ch_volume_sum),
            avg_mat = weighted_average(mat_cost_sum, mat_volume_sum)
        )

    # ---------------------------------------------------------------------- #

    def _create_statistics_sheet(
        self, wb: Workbook, category_stats: list[CategoryStats],
        unique_transports: int, unique_deposits: int
    ) -> None:
        """Creates the statistics sheet. category_stats is a 7-item list of category aggregates
        in column order: all, in, out, prestari in, prestari out, exploatare in, exploatare out."""
        ws = wb.create_sheet("Statistici", 1)
        NUM_STAT_COLS = 8 # Cols A–H: the main statistics table spans the whole sheet

        # ---------------------------------------------------------------------- #

        def write_block_header(row: int, text: str) -> None:
            """Helper: write a full-width light-blue block header row.
            Visually separates sections (VOLUME, PREȚURI, etc.) in the stats table"""
            self._apply_styles_to_cell(
                ws.cell(row = row, column = 1, value = text), self.stats_section_font, 
                self.light_blue_fill, self.center_align
            )
            for c in range(2, NUM_STAT_COLS + 1):
                self._apply_styles_to_cell(
                    ws.cell(row = row, column = c), self.calibri_12, 
                    self.light_blue_fill, self.center_align
                )

        # ---------------------------------------------------------------------- #

        def write_data_row(row: int, label: str, values: list, fmt: str | None) -> None:
            """Helper: write a data row — label in col A, 7 values in B-H"""
            self._apply_styles_to_cell(
                ws.cell(row = row, column = 1, value = label), self.calibri_12_bold, 
                self.light_blue_fill, self.center_align
            )
            for col_off, value in enumerate(values, start = 2):
                cell = ws.cell(row = row, column = col_off, value = value)
                nr_fmt = fmt if fmt and value != 0 and value != "" else None
                self._apply_styles_to_cell(cell, self.calibri_12, alignment = self.center_align, nr_format = nr_fmt)

        # ---------------------------------------------------------------------- #

        # Title banner: merged across A–H, 2 rows tall
        # Row 3 is left empty as a visual spacer; data starts at row 4
        ws.merge_cells(start_row = 1, start_column = 1, end_row = 2, end_column = NUM_STAT_COLS)

        self._apply_styles_to_cell(
            ws.cell(row = 1, column = 1, value = "STATISTICI GENERALE"),
            self.stats_font_header, self.stats_title_fill, self.center_align
        )

        current_row = 4  # column headers start at row 4

        # Write column headers — one per sheet category
        for col_idx, header in enumerate([
            "Categorie",
            "Total avize",
            "Total intrări",
            "Total ieșiri",
            "Prestări intrare",
            "Prestări ieșire",
            "Exploatare intrare",
            "Exploatare ieșire"
        ], start = 1):
            self._apply_styles_to_cell(
                ws.cell(row = current_row, column = col_idx, value = header),
                self.calibri_12_bold_white, self.dark_blue_fill, self.center_align
            )
        current_row += 1

        # ================================================================
        # BLOCK 1 — VOLUME (m³)
        # "Volum mediu/aviz" (last row) is total ÷ n at precision 2 since averaged
        # values don't need 6-decimal display; all others are totals at 6 decimals.
        # ================================================================
        write_block_header(current_row, "VOLUME (m³)")
        current_row += 1

        for label, attr, fmt in [
            ("Volum total materiale", "v_total_mat", self._FMT_VOLUME),
            ("Volum total Lemn rotund", "v_total_lr", self._FMT_VOLUME),
            ("Volum total Lemn foc", "v_total_lf", self._FMT_VOLUME),
            ("Volum total Cherestele", "v_total_ch", self._FMT_VOLUME),
            ("Volum mediu/aviz", "v_avg_aviz", self._FMT_PRICE)
        ]:
            write_data_row(current_row, label, [getattr(stats, attr) for stats in category_stats], fmt)
            current_row += 1

        current_row += 1  # spacer row

        # ================================================================
        # BLOCK 2 — PREȚURI (RON)
        # ================================================================
        write_block_header(current_row, "PREȚURI (RON)")
        current_row += 1

        for label, attr in [
            ("Preț total materiale", "p_total_m"),
            ("Preț total transport", "p_total_transp"),
            ("Preț total exploatare", "p_total_expl"),
            ("Preț total general", "p_total_aviz"),
            ("Valoare totală impozit", "v_total_impozit")
        ]:
            write_data_row(current_row, label, [getattr(stats, attr) for stats in category_stats], self._FMT_PRICE)
            current_row += 1

        current_row += 1  # spacer row

        # ================================================================
        # BLOCK 3 — PREȚURI MEDII PE SORTIMENT (RON/m³)
        # Volume-weighted average unit price per material type, plus a combined
        # materiale price. Computed in _aggregate_category_statistics.
        # ================================================================
        write_block_header(current_row, "PREȚURI MEDII PE SORTIMENT (RON / m³)")
        current_row += 1

        for label, attr in [
            ("Preț mediu Lemn rotund", "avg_lr"),
            ("Preț mediu Lemn foc", "avg_lf"),
            ("Preț mediu Cherestele", "avg_ch"),
            ("Preț mediu materiale", "avg_mat")
        ]:
            values = [getattr(stats, attr) for stats in category_stats]
            # The "all" column (index 0) blends intrare (buy) and iesire (sell) unit
            # prices into a figure that maps to no real per-m³ rate, so it's shown as 0.
            # Only the per-direction columns carry a meaningful average.
            values[0] = 0
            write_data_row(current_row, label, values, self._FMT_PRICE)
            current_row += 1

        current_row += 1  # spacer row

        # ================================================================
        # BLOCK 4 — ALTE STATISTICI
        # Transport and deposit counts only fill col B (Total avize);
        # the remaining 6 columns are left empty for those rows.
        # ================================================================
        write_block_header(current_row, "ALTE STATISTICI")
        current_row += 1

        for label, values in [
            ("Număr avize", [stats.notice_count for stats in category_stats]),
            ("Număr mijloace transport", [unique_transports, "", "", "", "", "", ""]),
            ("Număr depozite utilizate", [unique_deposits, "", "", "", "", "", ""])
        ]:
            write_data_row(current_row, label, values, '0')
            current_row += 1

        # Fixed column widths for the stats table (A–H)
        ws.column_dimensions['A'].width = 42
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 22

    # ---------------------------------------------------------------------- #

    def _aggregate_species_volumes(self, notice_list: list[TransportNoticeModel], attr: str) -> dict[str, float]:
        """Sum per-species volumes across a list of notices for one good type.
        attr: one of 'volume_pe_specii_lemn_rotund' / '_lemn_foc' / '_cherestele'.
        Per-notice species volumes are already snapped to 6 decimals in
        infer_volume_totals, but summing many of them reintroduces float noise
        below the 6th decimal — so we snap the final per-species total back to
        the source precision (6 decimals). Returns {species_name: total_m3}
        """
        result: dict[str, float] = {}

        for notice in notice_list:
            for specie, volume in getattr(notice.totals, attr).items():
                result[specie] = result.get(specie, 0.0) + volume

        return {specie: round(volume, VOLUME_PRECISION) for specie, volume in result.items()}

    # ---------------------------------------------------------------------- #

    def _create_totaluri_sheet(self, wb: Workbook, notices: list[TransportNoticeModel]) -> None:
        """Creates the 'Totaluri' sheet with volume-by-species tables for intrări and ieșiri.
        Each side stacks three material sections (Lemn rotund, Lemn foc, Cherestele),
        sorted by volume descending, with a TOTAL row per section
        """
        ws = wb.create_sheet("Totaluri", 2)

        # Column layout — two side-by-side 2-col tables with a narrow spacer.
        COL_IN = 1 # A-B: Intrări
        COL_SPACE = 3 # C: spacer
        COL_OUT = 4 # D-E: Ieșiri

        intrare_notices = [n for n in notices if n.type in self.INTRARE_T]
        iesire_notices = [n for n in notices if n.type in self.IESIRE_T]

        # ---------------------------------------------------------------------- #

        specii_intrari_lr = self._aggregate_species_volumes(intrare_notices, "volume_pe_specii_lemn_rotund")
        specii_intrari_lf = self._aggregate_species_volumes(intrare_notices, "volume_pe_specii_lemn_foc")
        specii_intrari_ch = self._aggregate_species_volumes(intrare_notices, "volume_pe_specii_cherestele")

        specii_iesiri_lr = self._aggregate_species_volumes(iesire_notices, "volume_pe_specii_lemn_rotund")
        specii_iesiri_lf = self._aggregate_species_volumes(iesire_notices, "volume_pe_specii_lemn_foc")
        specii_iesiri_ch = self._aggregate_species_volumes(iesire_notices, "volume_pe_specii_cherestele")

        self.logger.debug(
            f"Species aggregation (intrari): LR = {specii_intrari_lr} LF = {specii_intrari_lf} CH = {specii_intrari_ch}"
        )
        self.logger.debug(
            f"Species aggregation (iesiri): LR = {specii_iesiri_lr} LF = {specii_iesiri_lf} CH = {specii_iesiri_ch}"
        )

        # ---------------------------------------------------------------------- #

        def write_species_section(title: str, species_dict: dict[str, float], start_row: int, start_col: int) -> int:
            """Write one material block (e.g. LEMN ROTUND): header row + sorted species rows + TOTAL.
            Returns the row number of the TOTAL row so the caller can advance current_row."""
            row = start_row

            self._apply_styles_to_cell(
                ws.cell(row = row, column = start_col, value = title),
                self.calibri_12_bold_white, self.dark_blue_fill, self.center_align
            )
            self._apply_styles_to_cell(
                ws.cell(row = row, column = start_col + 1, value = "Total (m³)"),
                self.calibri_12_bold, self.light_blue_fill, self.center_align
            )

            row += 1
            for specie, volume in sorted(species_dict.items(), key = lambda entry: entry[1], reverse = True):
                for col, val in [(start_col, specie), (start_col + 1, volume)]:
                    cell = ws.cell(row = row, column = col, value = val)
                    nr_fmt = self._FMT_VOLUME if isinstance(val, float) else None
                    self._apply_styles_to_cell(cell, self.calibri_12, alignment = self.center_align, nr_format = nr_fmt)
                row += 1

            # Snap to 6 decimals to remove float-sum noise — species_dict values are
            # already 6-decimal-snapped, this just protects the final sum.
            total_vol = round(sum(species_dict.values()), VOLUME_PRECISION)
            for col, val in [(start_col, "TOTAL"), (start_col + 1, total_vol)]:
                cell = ws.cell(row = row, column = col, value = val)
                nr_fmt = self._FMT_VOLUME if isinstance(val, float) else None
                self._apply_styles_to_cell(cell, self.calibri_12_bold, self.light_blue_fill, self.center_align, nr_format = nr_fmt)
            return row

        # ---------------------------------------------------------------------- #

        # Intrări side — title banner spans 2 rows
        current_row = 1
        ws.merge_cells(
            start_row = current_row, start_column = COL_IN,
            end_row = current_row + 1, end_column = COL_IN + 1
        )

        self._apply_styles_to_cell(
            ws.cell(row = current_row, column = COL_IN, value = "VOLUME PE SPECII — INTRĂRI"),
            self.stats_font_header, self.stats_title_fill, self.center_align
        )

        current_row += 3  # skip 2 title rows + 1 spacer

        for species_dict, label in [(specii_intrari_lr, "LEMN ROTUND"),
                                    (specii_intrari_lf, "LEMN FOC"),
                                    (specii_intrari_ch, "CHERESTELE")]:

            total_row = write_species_section(label, species_dict, current_row, COL_IN)
            current_row = total_row + 2  # + 2 leaves one blank row between sections

        # Ieșiri side — same vertical layout, 3 columns to the right
        current_row = 1
        ws.merge_cells(
            start_row = current_row, start_column = COL_OUT,
            end_row = current_row + 1, end_column = COL_OUT + 1
        )

        self._apply_styles_to_cell(
            ws.cell(row = current_row, column = COL_OUT, value = "VOLUME PE SPECII — IEȘIRI"),
            self.stats_font_header, self.stats_title_fill, self.center_align
        )
        current_row += 3
        for species_dict, label in [(specii_iesiri_lr, "LEMN ROTUND"),
                                    (specii_iesiri_lf, "LEMN FOC"),
                                    (specii_iesiri_ch, "CHERESTELE")]:
            total_row = write_species_section(label, species_dict, current_row, COL_OUT)
            current_row = total_row + 2

        # Fixed column widths
        ws.column_dimensions[utils.get_column_letter(COL_IN)].width = 22
        ws.column_dimensions[utils.get_column_letter(COL_IN + 1)].width = 15
        ws.column_dimensions[utils.get_column_letter(COL_SPACE)].width = 4
        ws.column_dimensions[utils.get_column_letter(COL_OUT)].width = 22
        ws.column_dimensions[utils.get_column_letter(COL_OUT + 1)].width = 15

        # ================================================================
        # TOTALURI PE SORTIMENTE
        # Placed to the right of the species tables at col G (col 7).
        # ================================================================
        SPACE_COL = 6
        TOTALS_COL = 7
        TOTALSWIDTH = 4 # Sortiment | Intrări | Ieșiri | Total

        totals = [
            ("LEMN ROTUND",
             round(sum(specii_intrari_lr.values()), VOLUME_PRECISION),
             round(sum(specii_iesiri_lr.values()), VOLUME_PRECISION)
            ),
            ("LEMN FOC",
             round(sum(specii_intrari_lf.values()), VOLUME_PRECISION),
             round(sum(specii_iesiri_lf.values()), VOLUME_PRECISION)
            ),
            ("CHERESTELE",
             round(sum(specii_intrari_ch.values()), VOLUME_PRECISION),
             round(sum(specii_iesiri_ch.values()), VOLUME_PRECISION)
            )
        ]

        # Title banner
        ws.merge_cells(
            start_row = 1, start_column = TOTALS_COL,
            end_row = 2, end_column = TOTALS_COL + TOTALSWIDTH - 1
        )

        self._apply_styles_to_cell(
            ws.cell(row = 1, column = TOTALS_COL, value = "TOTALURI PE SORTIMENTE (m³)"),
            self.stats_font_header, self.stats_title_fill, self.center_align
        )

        HEADER_ROW = 4
        for offset, header in enumerate(["Sortiment", "Intrări", "Ieșiri", "Total"]):
            self._apply_styles_to_cell(
                ws.cell(row = HEADER_ROW, column = TOTALS_COL + offset, value = header),
                self.calibri_12_bold_white, self.dark_blue_fill, self.center_align
            )

        DATA_START = HEADER_ROW + 1
        for row_idx, (label, total_intrari, total_iesiri) in enumerate(totals, start = DATA_START):
            balance = round(total_intrari - total_iesiri, VOLUME_PRECISION)
            self._apply_styles_to_cell(
                ws.cell(row = row_idx, column = TOTALS_COL, value = label),
                self.calibri_12_bold, self.light_blue_fill, self.center_align
            )
            for offset, val in enumerate([total_intrari, total_iesiri, balance], start = 1):
                cell = ws.cell(row = row_idx, column = TOTALS_COL + offset, value = val)
                self._apply_styles_to_cell(cell, self.calibri_12, alignment = self.center_align, nr_format = self._FMT_VOLUME)

        # Column widths
        ws.column_dimensions[utils.get_column_letter(SPACE_COL)].width = 4
        ws.column_dimensions[utils.get_column_letter(TOTALS_COL)].width = 30
        for i in range(1, TOTALSWIDTH):
            ws.column_dimensions[utils.get_column_letter(TOTALS_COL + i)].width = 18

    # ---------------------------------------------------------------------- #

    def generate_report(
            self, sorted_notices: list[TransportNoticeModel], deposit_data: list[DepozitDataModel],
            output_path: str, prestari_notices: list[TransportNoticeModel]
        ) -> None:
        """Driver code for report generation, invoked by main window when the user presses generate report button"""
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Compute every notice's totals once — these are the single source of truth
        # for all sheets (rows, totals, statistics) and the validation JSON.
        for notice in sorted_notices:
            self._populate_notice_totals(notice, self._get_deposit_for_notice(notice, deposit_data))

        # Create the Date depozite sheet where the user prices are displayed
        self._create_deposit_data_sheet(wb, deposit_data)

        # Filter notices into the categories that become individual sheets and the
        # statistics columns.
        notices_all = sorted_notices
        
        notices_in = [n for n in sorted_notices if n.type in self.INTRARE_T]
        notices_out = [n for n in sorted_notices if n.type in self.IESIRE_T]

        notices_pin = [n for n in prestari_notices if n.type in self.INTRARE_T]
        notices_pout = [n for n in prestari_notices if n.type in self.IESIRE_T]

        notices_ein = [n for n in sorted_notices if n.type == NoticeType.INTRARE_DIN_PARTIDA_PROPRIE]
        notices_eout = [n for n in sorted_notices if n.type == NoticeType.IESIRE_DIN_DEPOZIT_LR]

        self._make_notice_sheet(wb, self.SH_ALL, notices_all, "TabelToateAvizele")

        self._make_notice_sheet(wb, self.SH_IN, notices_in, "TabelIntrari")
        self._make_notice_sheet(wb, self.SH_OUT, notices_out, "TabelIesiri")

        self._make_notice_sheet(wb, self.SH_PIN, notices_pin, "TabelPrestariIntrare")
        self._make_notice_sheet(wb, self.SH_POUT, notices_pout, "TabelPrestariIesire")
        
        self._make_notice_sheet(wb, self.SH_EIN, notices_ein, "TabelExploatareIntrare")
        self._make_notice_sheet(wb, self.SH_EOUT, notices_eout, "TabelExploatareIesire")

        self.logger.debug(
            "Sheet notice counts:\n"
            f" All notices:        {len(notices_all)}\n"
            f" Intrare:            {len(notices_in)}\n"
            f" Iesire:             {len(notices_out)}\n"
            f" Prestari intrare:   {len(notices_pin)}\n"
            f" Prestari iesire:    {len(notices_pout)}\n"
            f" Exploatare intrare: {len(notices_ein)}\n"
            f" Exploatare iesire:  {len(notices_eout)}"
        )

        # All car numbers set
        all_transports: set[str] = {
            n.cap_tractor.strip()
            for n in sorted_notices
            if n.cap_tractor.strip()
        }

        # LIR car numbers - 1 sheet for 2 car numbers
        lir_transports: set[str] = {t for t in all_transports if "LIR" in t.upper()}

        # Other car numbers - 1 sheet for each
        other_transports: set[str] = all_transports - lir_transports

        self.logger.debug(
            f"Transport sheets: {len(all_transports)} unique transports "
            f"({len(lir_transports)} LIR, {len(other_transports)} other)"
        )

        if lir_transports:
            self._create_transport_sheet(wb, sorted_notices, "Avize_LIR", transport_filter = None)
        for transport in sorted(other_transports):
            self._create_transport_sheet(wb, sorted_notices, f"Avize_{transport}", transport_filter = transport)

        # all_transports was already computed above for the transport sheets — reuse it.
        unique_transports: int  = len(all_transports)
        unique_deposits: int = len({
            n.provenienta.strip()
            for n in sorted_notices if n.provenienta.strip()
        })

        category_stats: list[CategoryStats] = [
            self._aggregate_category_statistics(category_notices)
            for category_notices in (notices_all, notices_in, notices_out,
                                     notices_pin, notices_pout, notices_ein, notices_eout)
        ]

        self.logger.debug(
            f"Statistics: all = {category_stats[0].notice_count}, in = {category_stats[1].notice_count}, out = {category_stats[2].notice_count}, "
            f"p_in = {category_stats[3].notice_count}, p_out = {category_stats[4].notice_count}, e_in = {category_stats[5].notice_count}, e_out = {category_stats[6].notice_count}, "
            f"transports = {unique_transports}, deposits = {unique_deposits}"
        )

        self._create_statistics_sheet(wb, category_stats, unique_transports, unique_deposits)

        self._create_totaluri_sheet(wb, sorted_notices)

        wb.save(output_path)
