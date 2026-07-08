import openpyxl
import pytest
from unittest.mock import MagicMock
from core.report_generator import ReportGenerator, CategoryStats
from core.enums import NoticeType, DepozitSource
from tests.conftest import make_notice, make_wood_item, make_deposit_data

@pytest.fixture
def rg():
    """ReportGenerator with a mocked logger."""
    return ReportGenerator(MagicMock())

# ---------------------------------------------------------------------- #

def _populated_notice(rg, notice_type, lr_vol = 3.0, lf_vol = 2.0, ch_vol = 0.0, **dd_prices):
    """Build a notice with volumes + type, then populate its totals from a deposit."""
    total = lr_vol + lf_vol + ch_vol
    notice = make_notice(volum_total_aviz = total)
    items = []
    if lr_vol > 0:
        items.append(make_wood_item(sortiment = "Lemn rotund", volum_mc = lr_vol, specie = "Brad"))
    if lf_vol > 0:
        items.append(make_wood_item(sortiment = "Lemn de foc", volum_mc = lf_vol, specie = "Fag", nr_crt = 2))
    if ch_vol > 0:
        items.append(make_wood_item(sortiment = "Cherestele", volum_mc = ch_vol, specie = "Stejar", nr_crt = 3))
    notice.wood_items = items
    notice.infer_volume_totals()
    notice.type = notice_type
    dd = make_deposit_data(tip = "Depozit extern", sursa = DepozitSource.AVIZE, **dd_prices)
    rg._populate_notice_totals(notice, dd)
    return notice

# ---------------------------------------------------------------------- #

class TestPopulateNoticeTotals:

    def _setup_notice(self, notice_type, lr_vol = 3.0, lf_vol = 2.0, ch_vol = 0.0):
        """Helper: create a notice with pre-set volumes and type."""
        total = lr_vol + lf_vol + ch_vol
        notice = make_notice(volum_total_aviz = total)
        items = []
        if lr_vol > 0:
            items.append(make_wood_item(sortiment = "Lemn rotund", volum_mc = lr_vol, specie = "Brad"))
        if lf_vol > 0:
            items.append(make_wood_item(sortiment = "Lemn de foc", volum_mc = lf_vol, specie = "Fag", nr_crt = 2))
        if ch_vol > 0:
            items.append(make_wood_item(sortiment = "Cherestele", volum_mc = ch_vol, specie = "Stejar", nr_crt = 3))
        notice.wood_items = items
        notice.infer_volume_totals()
        notice.type = notice_type
        return notice

    # ---------------------------------------------------------------------- #

    def test_intrare_uses_intrare_prices(self, rg):
        notice = self._setup_notice(NoticeType.INTRARE_DIN_SURSA_EXTERNA)
        dd = make_deposit_data(
            tip = "Depozit extern", sursa = DepozitSource.AVIZE,
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50, p_intrare_ch_mc = 0, p_transport_mc = 30
        )
        rg._populate_notice_totals(notice, dd)

        assert notice.totals.pret_mc_lemn_rotund == 100  # intrare price
        assert notice.totals.pret_mc_lemn_foc == 50
        assert notice.totals.pret_total_materiale == (3.0 * 100) + (2.0 * 50)
        assert notice.totals.pret_total_transport == 5.0 * 30
        assert notice.totals.pret_total_exploatare == 0.0  # no exploatare for external
        assert notice.totals.valoare_impozit_punere_in_piata == 0.0

    # ---------------------------------------------------------------------- #

    def test_iesire_uses_iesire_prices(self, rg):
        notice = self._setup_notice(NoticeType.IESIRE_DIN_DEPOZIT_PRINCIPAL)
        dd = make_deposit_data(
            name = "VALENII DE MURES 95", tip = "Depozit",
            p_iesire_lr_mc = 200, p_iesire_lf_mc = 80, p_iesire_ch_mc = 0, p_transport_mc = 40
        )
        rg._populate_notice_totals(notice, dd)

        assert notice.totals.pret_mc_lemn_rotund == 200  # iesire price
        assert notice.totals.pret_mc_lemn_foc == 80
        assert notice.totals.pret_total_exploatare == 0.0

    # ---------------------------------------------------------------------- #

    def test_intrare_partida_proprie_has_exploatare_and_impozit(self, rg):
        notice = self._setup_notice(NoticeType.INTRARE_DIN_PARTIDA_PROPRIE)
        dd = make_deposit_data(
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50, p_transport_mc = 30, p_expl_mc = 20
        )
        rg._populate_notice_totals(notice, dd)

        pret_mat = (3.0 * 100) + (2.0 * 50)  # 400
        pret_expl = 5.0 * 20  # 100
        expected_impozit = (pret_mat + pret_expl) * 0.005

        assert notice.totals.pret_total_exploatare == pret_expl
        assert notice.totals.valoare_impozit_punere_in_piata == pytest.approx(expected_impozit)
        assert notice.totals.pret_total_aviz == pret_mat + (5.0 * 30) + pret_expl

    # ---------------------------------------------------------------------- #

    def test_iesire_lr_has_exploatare_and_different_impozit(self, rg):
        notice = self._setup_notice(NoticeType.IESIRE_DIN_DEPOZIT_LR)
        dd = make_deposit_data(
            p_iesire_lr_mc = 200, p_iesire_lf_mc = 80, p_transport_mc = 40, p_expl_mc = 25
        )
        rg._populate_notice_totals(notice, dd)

        pret_mat = (3.0 * 200) + (2.0 * 80)  # 760
        expected_impozit = pret_mat * 0.005  # only materiale, no exploatare

        assert notice.totals.pret_total_exploatare == 5.0 * 25
        assert notice.totals.valoare_impozit_punere_in_piata == pytest.approx(expected_impozit)

# ---------------------------------------------------------------------- #

class TestGetDepositForNotice:

    def test_finds_matching_deposit(self, rg):
        notice = make_notice(provenienta = "Depozit A")
        deposits = [
            make_deposit_data(name = "Depozit B"),
            make_deposit_data(name = "Depozit A"),
        ]
        result = rg._get_deposit_for_notice(notice, deposits)
        assert result.nume_depozit == "Depozit A"

    # ---------------------------------------------------------------------- #

    def test_returns_none_on_missing_deposit(self, rg):
        notice = make_notice(provenienta = "Inexistent")
        deposits = [make_deposit_data(name = "Altceva")]
        assert rg._get_deposit_for_notice(notice, deposits) is None

# ---------------------------------------------------------------------- #

class TestAggregateCategory:
    """Statistics are computed in Python (no Excel formulas). These cover the math
    that the SUM / SUMPRODUCT formulas used to do."""

    def test_basic_sums_and_weighted_prices(self, rg):
        # 3 m³ LR @ 100, 2 m³ LF @ 50, transport 30/m³
        notice = _populated_notice(
            rg, NoticeType.INTRARE_DIN_SURSA_EXTERNA,
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50, p_transport_mc = 30
        )
        agg = rg._aggregate_category_statistics([notice])

        assert isinstance(agg, CategoryStats)
        assert agg.notice_count == 1
        assert agg.v_total_mat == 5.0
        assert agg.v_total_lr == 3.0
        assert agg.v_total_lf == 2.0
        assert agg.v_total_ch == 0.0
        assert agg.p_total_m == (3.0 * 100) + (2.0 * 50)   # 400
        assert agg.p_total_transp == 5.0 * 30              # 150
        assert agg.v_avg_aviz == 5.0
        # Volume-weighted unit prices (single notice -> the unit price itself)
        assert agg.avg_lr == 100
        assert agg.avg_lf == 50
        assert agg.avg_ch == 0.0
        # Combined materiale price = total materiale / total volume = 400 / 5
        assert agg.avg_mat == 80.0

    # ---------------------------------------------------------------------- #

    def test_weighted_price_ignores_zero_volume_notices(self, rg):
        # One notice has LR volume, another has none — the empty one must not drag
        # the weighted average toward 0.
        n1 = _populated_notice(rg, NoticeType.INTRARE_DIN_SURSA_EXTERNA,
                               lr_vol = 4.0, lf_vol = 0.0, p_intrare_lr_mc = 100)
        n2 = _populated_notice(rg, NoticeType.INTRARE_DIN_SURSA_EXTERNA,
                               lr_vol = 0.0, lf_vol = 3.0, p_intrare_lr_mc = 999, p_intrare_lf_mc = 50)
        agg = rg._aggregate_category_statistics([n1, n2])

        assert agg.avg_lr == 100   # only n1 has LR volume
        assert agg.v_total_lr == 4.0

    # ---------------------------------------------------------------------- #

    def test_empty_category_is_all_zero(self, rg):
        agg = rg._aggregate_category_statistics([])
        assert agg.notice_count == 0
        assert agg.v_total_mat == 0.0
        assert agg.p_total_m == 0.0
        assert agg.v_avg_aviz == 0.0
        assert agg.avg_lr == 0.0
        assert agg.avg_mat == 0.0

# ---------------------------------------------------------------------- #

class TestGenerateReportEmitsValues:
    """End-to-end: data cells are static numbers; the only formulas allowed are the
    per-column SUM totals in the TOTAL row."""

    def test_cells_are_computed_values_not_formulas(self, rg, tmp_path):
        notice = _populated_notice(
            rg, NoticeType.INTRARE_DIN_SURSA_EXTERNA,
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50, p_transport_mc = 30
        )
        deposit = make_deposit_data(
            name = notice.provenienta, tip = "Depozit extern", sursa = DepozitSource.AVIZE,
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50, p_transport_mc = 30
        )
        out = tmp_path / "report.xlsx"
        rg.generate_report([notice], [deposit], str(out), [])

        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert {"Date depozite", "Statistici", "Totaluri", "Toate avizele"} <= set(wb.sheetnames)

        ws = wb["Toate avizele"]
        # Column K (11) = Preț total materiale on the first data row
        cell = ws.cell(row = 2, column = 11)
        assert isinstance(cell.value, (int, float))
        assert not (isinstance(cell.value, str) and cell.value.startswith("="))
        assert cell.value == 400.0

        # TOTAL row is a filter-aware SUBTOTAL over the (single) data row
        total_cell = ws.cell(row = 3, column = 11)
        assert total_cell.value == "=SUBTOTAL(109, K2:K2)"

    # ---------------------------------------------------------------------- #

    def test_only_total_rows_use_sum_formulas(self, rg, tmp_path):
        """The TOTAL rows SUBTOTAL their column; no other formula type (SUMPRODUCT,
        cross-sheet references, per-row price formulas) may reappear anywhere."""
        notice = _populated_notice(
            rg, NoticeType.IESIRE_DIN_DEPOZIT_LR,
            p_iesire_lr_mc = 200, p_iesire_lf_mc = 80, p_transport_mc = 40, p_expl_mc = 25
        )
        deposit = make_deposit_data(
            name = notice.provenienta, tip = "Depozit Temporar LR",
            p_iesire_lr_mc = 200, p_iesire_lf_mc = 80, p_transport_mc = 40, p_expl_mc = 25
        )
        out = tmp_path / "report.xlsx"
        rg.generate_report([notice], [deposit], str(out), [])

        wb = openpyxl.load_workbook(out)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        assert cell.value.startswith("=SUBTOTAL("), \
                            f"unexpected formula in {ws.title}!{cell.coordinate}: {cell.value}"

# ---------------------------------------------------------------------- #

class TestAggregateSpeciesVolumes:
    """Cross-notice species sums must stay snapped to 6-decimal source precision."""

    def _notice_with_species(self, species_volumes: dict[str, float]):
        """Build a notice whose totals.volume_pe_specii_lemn_rotund holds the given dict."""
        notice = make_notice()
        notice.totals.volume_pe_specii_lemn_rotund = dict(species_volumes)
        return notice

    # ---------------------------------------------------------------------- #

    def test_empty_notice_list_returns_empty_dict(self, rg):
        assert rg._aggregate_species_volumes([], "volume_pe_specii_lemn_rotund") == {}

    # ---------------------------------------------------------------------- #

    def test_single_notice_passes_through(self, rg):
        n = self._notice_with_species({"Brad": 1.5, "Fag": 2.25})
        result = rg._aggregate_species_volumes([n], "volume_pe_specii_lemn_rotund")
        assert result == {"Brad": 1.5, "Fag": 2.25}

    # ---------------------------------------------------------------------- #

    def test_sums_same_species_across_notices(self, rg):
        notices = [
            self._notice_with_species({"Brad": 1.0, "Fag": 2.0}),
            self._notice_with_species({"Brad": 3.0, "Stejar": 4.5}),
            self._notice_with_species({"Brad": 0.5}),
        ]
        result = rg._aggregate_species_volumes(notices, "volume_pe_specii_lemn_rotund")
        assert result == {"Brad": 4.5, "Fag": 2.0, "Stejar": 4.5}

    # ---------------------------------------------------------------------- #

    def test_float_noise_snapped_at_6_decimals(self, rg):
        """0.1 + 0.2 == 0.30000000000000004 — must be snapped to 0.3."""
        notices = [
            self._notice_with_species({"Brad": 0.1}),
            self._notice_with_species({"Brad": 0.2}),
        ]
        result = rg._aggregate_species_volumes(notices, "volume_pe_specii_lemn_rotund")
        assert result == {"Brad": 0.3}

    # ---------------------------------------------------------------------- #

    def test_six_decimal_source_values_sum_clean(self, rg):
        notices = [
            self._notice_with_species({"Fag": 0.123456}),
            self._notice_with_species({"Fag": 0.234567}),
            self._notice_with_species({"Fag": 0.111111}),
        ]
        result = rg._aggregate_species_volumes(notices, "volume_pe_specii_lemn_rotund")
        assert result == {"Fag": 0.469134}

    # ---------------------------------------------------------------------- #

    def test_unrelated_attr_is_isolated(self, rg):
        n = make_notice()
        n.totals.volume_pe_specii_lemn_rotund = {"Brad": 1.0}
        n.totals.volume_pe_specii_lemn_foc = {"Fag": 2.0}
        assert rg._aggregate_species_volumes([n], "volume_pe_specii_lemn_foc") == {"Fag": 2.0}

# ---------------------------------------------------------------------- #

class TestEmptyCategorySheet:

    def test_empty_sheet_shows_placeholder_and_no_total_row(self, rg, tmp_path):
        """A category with no notices (e.g. no prestări marked) must render a
        placeholder message instead of an empty table with a TOTAL row."""
        notice = _populated_notice(
            rg, NoticeType.INTRARE_DIN_SURSA_EXTERNA,
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50, p_transport_mc = 30
        )
        deposit = make_deposit_data(
            name = notice.provenienta, tip = "Depozit extern", sursa = DepozitSource.AVIZE,
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50, p_transport_mc = 30
        )
        out = tmp_path / "report.xlsx"
        rg.generate_report([notice], [deposit], str(out), [])  # no prestări

        wb = openpyxl.load_workbook(str(out))
        ws = wb["Avize prestari intrare"]
        assert ws.cell(row = 2, column = 1).value == "Nici un aviz de afișat"
        assert ws.cell(row = 3, column = 1).value is None  # no TOTAL row
