"""End-to-end regression test against the real month of SUMAL data in inputs/.

Runs the exact pipeline the app runs (folder validation -> parse -> deposit
init -> report generation) and cross-checks every computed number against an
independent recomputation done with Decimal arithmetic straight from the raw
CSV text — so any precision loss or aggregation bug shows up as an exact
mismatch, not a tolerance failure.
"""
import csv
import os
from decimal import Decimal

import openpyxl
import pytest
from unittest.mock import MagicMock

from core.pipeline import ReportPipeline
from core.config import DEPOSIT_DATA_ENABLED_FIELDS_BY_TYPE
from core.enums import FolderStatus, NoticeType
from core.utils import extract_base_name

INPUTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "inputs")

pytestmark = pytest.mark.skipif(
    not os.path.isdir(INPUTS_DIR), reason = "inputs/ folder with real SUMAL data not present"
)

INTRARE_TYPES = {NoticeType.INTRARE_DIN_PARTIDA_PROPRIE, NoticeType.INTRARE_DIN_SURSA_EXTERNA}
EXPLOATARE_TYPES = {NoticeType.INTRARE_DIN_PARTIDA_PROPRIE, NoticeType.IESIRE_DIN_DEPOZIT_LR}

# ---------------------------------------------------------------------- #

def read_raw_csv_volumes() -> dict[str, dict]:
    """Independently recompute per-notice volumes from the raw CSV text using
    Decimal, keyed by cod_unic. Duplicate files (same base name) are skipped
    the same way the app skips them."""
    raw: dict[str, dict] = {}
    seen_base_names: set[str] = set()
    for filename in sorted(os.listdir(INPUTS_DIR)):
        if not filename.lower().endswith(".csv"):
            continue
        base_name = extract_base_name(filename)
        if base_name in seen_base_names:
            continue
        seen_base_names.add(base_name)

        with open(os.path.join(INPUTS_DIR, filename), newline = "", encoding = "utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))

        entry = {
            "volum_total_aviz": Decimal(rows[0]["Volum total aviz"]),
            "per_sortiment": {},
            "per_species": {},
        }
        for row in rows:
            volum = Decimal(row["Volum(mc)"])
            sortiment = row["Sortiment"]
            entry["per_sortiment"][sortiment] = entry["per_sortiment"].get(sortiment, Decimal(0)) + volum
            key = (sortiment, row["Specie"])
            entry["per_species"][key] = entry["per_species"].get(key, Decimal(0)) + volum
        raw[rows[0]["Cod unic"]] = entry
    return raw

# ---------------------------------------------------------------------- #

def as_decimal(value: float) -> Decimal:
    """Exact Decimal of a float's shortest repr — a 6-decimal source value that
    survived parsing intact converts back to the identical Decimal."""
    return Decimal(repr(value))

# ---------------------------------------------------------------------- #

@pytest.fixture(scope = "module")
def pipeline(tmp_path_factory):
    """Run the full pipeline once for the whole module: parse, fill deterministic
    prices, generate the report, and return everything needed for assertions."""
    ctrl = ReportPipeline(MagicMock())
    ctrl.set_folder(INPUTS_DIR)
    status, _ = ctrl.validate_folder_files()
    assert status == FolderStatus.READY

    ctrl.parse_files()
    ctrl.initialize_deposit_data()

    # Deterministic unique price per (deposit, field) so price math is checkable.
    for deposit_idx, deposit in enumerate(ctrl.deposit_data):
        enabled_fields = DEPOSIT_DATA_ENABLED_FIELDS_BY_TYPE[deposit.tip_depozit]
        for field_idx, field_name in enumerate(sorted(enabled_fields)):
            setattr(deposit.price_data, field_name, 100 + deposit_idx * 10 + field_idx)

    assert ctrl.can_generate_reports()

    report_path = tmp_path_factory.mktemp("report") / "raport.xlsx"
    prestari = [n for n in ctrl.notices if n.type != NoticeType.INTRARE_DIN_PARTIDA_PROPRIE][:2]
    ctrl.generate_report(str(report_path), prestari)

    return ctrl, read_raw_csv_volumes(), openpyxl.load_workbook(str(report_path)), prestari

# ---------------------------------------------------------------------- #

class TestParsedNotices:

    def test_all_unique_notices_parsed_and_classified(self, pipeline):
        ctrl, raw, _, _ = pipeline
        assert len(ctrl.notices) == len(raw)
        assert {n.cod_unic for n in ctrl.notices} == set(raw)
        assert all(n.type != NoticeType.UNKNOWN for n in ctrl.notices)

    # ---------------------------------------------------------------------- #

    def test_notices_sorted_by_emission_date(self, pipeline):
        ctrl, _, _, _ = pipeline
        dates = [n.data_ora_emitere for n in ctrl.notices]
        assert dates == sorted(dates)

    # ---------------------------------------------------------------------- #

    def test_volume_totals_exact_at_source_precision(self, pipeline):
        """Every per-notice volume total must equal the Decimal sum of the raw
        CSV values exactly — no float drift at any of the 6 decimals."""
        ctrl, raw, _, _ = pipeline
        for notice in ctrl.notices:
            expected = raw[notice.cod_unic]
            assert as_decimal(notice.volum_total_aviz) == expected["volum_total_aviz"]
            assert as_decimal(notice.totals.volum_total_lemn_rotund) == \
                   expected["per_sortiment"].get("Lemn rotund", Decimal(0))
            assert as_decimal(notice.totals.volum_total_lemn_foc) == \
                   expected["per_sortiment"].get("Lemn de foc", Decimal(0))
            assert as_decimal(notice.totals.volum_total_cherestele) == \
                   expected["per_sortiment"].get("Cherestele", Decimal(0))

    # ---------------------------------------------------------------------- #

    def test_species_volumes_exact_at_source_precision(self, pipeline):
        ctrl, raw, _, _ = pipeline
        attr_by_sortiment = {
            "Lemn rotund": "volume_pe_specii_lemn_rotund",
            "Lemn de foc": "volume_pe_specii_lemn_foc",
            "Cherestele": "volume_pe_specii_cherestele",
        }
        for notice in ctrl.notices:
            for (sortiment, specie), volum in raw[notice.cod_unic]["per_species"].items():
                model_dict = getattr(notice.totals, attr_by_sortiment[sortiment])
                assert as_decimal(model_dict[specie]) == volum

# ---------------------------------------------------------------------- #

class TestPriceMath:

    def test_notice_prices_match_decimal_recompute(self, pipeline):
        """Recompute every notice's price totals with Decimal and compare within
        float-multiplication noise (far below the displayed 2 decimals)."""
        ctrl, raw, _, _ = pipeline
        tolerance = Decimal("1e-9")
        for notice in ctrl.notices:
            expected = raw[notice.cod_unic]
            deposit = next(d for d in ctrl.deposit_data if d.nume_depozit == notice.provenienta)
            prices = deposit.price_data
            if notice.type in INTRARE_TYPES:
                p_lr, p_lf, p_ch = prices.p_intrare_lr_mc, prices.p_intrare_lf_mc, prices.p_intrare_ch_mc
            else:
                p_lr, p_lf, p_ch = prices.p_iesire_lr_mc, prices.p_iesire_lf_mc, prices.p_iesire_ch_mc

            materiale = (
                expected["per_sortiment"].get("Lemn rotund", Decimal(0)) * p_lr
                + expected["per_sortiment"].get("Lemn de foc", Decimal(0)) * p_lf
                + expected["per_sortiment"].get("Cherestele", Decimal(0)) * p_ch
            )
            transport = expected["volum_total_aviz"] * prices.p_transport_mc
            exploatare = (
                expected["volum_total_aviz"] * prices.p_expl_mc
                if notice.type in EXPLOATARE_TYPES else Decimal(0)
            )

            assert abs(as_decimal(notice.totals.pret_total_materiale) - materiale) <= tolerance
            assert abs(as_decimal(notice.totals.pret_total_transport) - transport) <= tolerance
            assert abs(as_decimal(notice.totals.pret_total_exploatare) - exploatare) <= tolerance
            assert abs(
                as_decimal(notice.totals.pret_total_aviz) - (materiale + transport + exploatare)
            ) <= tolerance

    # ---------------------------------------------------------------------- #

    def test_impozit_only_on_partida_proprie_and_iesire_lr(self, pipeline):
        ctrl, _, _, _ = pipeline
        for notice in ctrl.notices:
            has_impozit = notice.totals.valoare_impozit_punere_in_piata > 0
            assert has_impozit == (notice.type in EXPLOATARE_TYPES)

# ---------------------------------------------------------------------- #

class TestGeneratedReport:

    def test_sheet_structure(self, pipeline):
        _, _, workbook, _ = pipeline
        expected_sheets = {
            "Date depozite", "Statistici", "Totaluri", "Toate avizele",
            "Avize intrare", "Avize iesire", "Avize prestari intrare",
            "Avize prestari iesire", "Avize exploatare intrare", "Avize exploatare iesire",
        }
        assert expected_sheets <= set(workbook.sheetnames)

    # ---------------------------------------------------------------------- #

    def test_toate_avizele_rows_match_notices_exactly(self, pipeline):
        ctrl, raw, workbook, _ = pipeline
        ws = workbook["Toate avizele"]
        codes_in_sheet = [
            ws.cell(row = row_idx, column = 2).value
            for row_idx in range(2, len(ctrl.notices) + 2)
        ]
        assert codes_in_sheet == [n.cod_unic for n in ctrl.notices]
        # Volume column (7) carries the exact source value for every row
        for row_idx, notice in enumerate(ctrl.notices, start = 2):
            cell_value = ws.cell(row = row_idx, column = 7).value
            assert as_decimal(cell_value) == raw[notice.cod_unic]["volum_total_aviz"]

    # ---------------------------------------------------------------------- #

    def test_intrare_iesire_sheets_partition_all_notices(self, pipeline):
        ctrl, _, workbook, _ = pipeline
        intrare_codes = {n.cod_unic for n in ctrl.notices if n.type in INTRARE_TYPES}
        iesire_codes = {n.cod_unic for n in ctrl.notices if n.type not in INTRARE_TYPES}

        def sheet_codes(sheet_name, count):
            ws = workbook[sheet_name]
            return {ws.cell(row = r, column = 2).value for r in range(2, count + 2)}

        assert sheet_codes("Avize intrare", len(intrare_codes)) == intrare_codes
        assert sheet_codes("Avize iesire", len(iesire_codes)) == iesire_codes

    # ---------------------------------------------------------------------- #

    def test_prestari_notices_stay_in_normal_flow(self, pipeline):
        """Prestări sheets are a filtered view — marked notices must ALSO remain
        in the main in/out sheets (domain rule)."""
        ctrl, _, workbook, prestari = pipeline
        assert prestari, "fixture must mark at least one prestări notice"

        prestari_in = {n.cod_unic for n in prestari if n.type in INTRARE_TYPES}
        prestari_out = {n.cod_unic for n in prestari if n.type not in INTRARE_TYPES}

        def sheet_codes(sheet_name):
            ws = workbook[sheet_name]
            codes = set()
            for row in ws.iter_rows(min_row = 2, max_col = 2):
                value = row[1].value
                if value:
                    codes.add(value)
            return codes

        assert prestari_in <= sheet_codes("Avize prestari intrare")
        assert prestari_out <= sheet_codes("Avize prestari iesire")
        # ...and still present in the unfiltered flow
        assert prestari_in <= sheet_codes("Avize intrare")
        assert prestari_out <= sheet_codes("Avize iesire")

    # ---------------------------------------------------------------------- #

    def test_statistici_volume_totals_match_decimal_sums(self, pipeline):
        """The Statistici sheet's volume rows (rows 6-9, cols B/C/D = all/in/out)
        must equal the Decimal grand totals of the source data."""
        ctrl, raw, workbook, _ = pipeline
        ws = workbook["Statistici"]

        def expected_totals(codes):
            totals = {"aviz": Decimal(0), "Lemn rotund": Decimal(0),
                      "Lemn de foc": Decimal(0), "Cherestele": Decimal(0)}
            for cod in codes:
                totals["aviz"] += raw[cod]["volum_total_aviz"]
                for sortiment in ("Lemn rotund", "Lemn de foc", "Cherestele"):
                    totals[sortiment] += raw[cod]["per_sortiment"].get(sortiment, Decimal(0))
            return totals

        categories = {
            2: expected_totals(list(raw)),  # col B: all
            3: expected_totals([n.cod_unic for n in ctrl.notices if n.type in INTRARE_TYPES]),
            4: expected_totals([n.cod_unic for n in ctrl.notices if n.type not in INTRARE_TYPES]),
        }
        row_by_label = {"aviz": 6, "Lemn rotund": 7, "Lemn de foc": 8, "Cherestele": 9}
        for col_idx, totals in categories.items():
            for label, row_idx in row_by_label.items():
                assert as_decimal(ws.cell(row = row_idx, column = col_idx).value) == totals[label], \
                    f"Statistici row {row_idx} col {col_idx} ({label})"

    # ---------------------------------------------------------------------- #

    def test_transport_sheets_cover_every_notice(self, pipeline):
        """Each notice appears in exactly one transport sheet: the shared LIR
        sheet if its plate contains LIR, otherwise its own plate sheet."""
        ctrl, _, workbook, _ = pipeline
        transport_sheets = [s for s in workbook.sheetnames if s.startswith("Avize_")]

        seen_codes: list[str] = []
        for sheet_name in transport_sheets:
            ws = workbook[sheet_name]
            for row in ws.iter_rows(min_row = 2, max_col = 6):
                cod, plate = row[1].value, row[5].value
                if cod is None or cod == "TOTAL" or cod == "Nici un aviz de afișat":
                    continue
                seen_codes.append(cod)
                if sheet_name == "Avize_LIR":
                    assert "LIR" in plate.upper()
                else:
                    assert plate.strip() == sheet_name.removeprefix("Avize_")

        assert sorted(seen_codes) == sorted(n.cod_unic for n in ctrl.notices)
