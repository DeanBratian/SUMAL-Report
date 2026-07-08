import pytest
import os
import tempfile
from unittest.mock import MagicMock
from openpyxl import Workbook
from core.excel_reader import ExcelReader
from core.sr_error import SRError

@pytest.fixture
def reader():
    return ExcelReader(MagicMock())

# ---------------------------------------------------------------------- #

def _create_xlsx(tmp_dir, filename, headers, rows):
    """Create a minimal .xlsx file with the given headers and row data."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path = os.path.join(tmp_dir, filename)
    wb.save(path)
    return path

# ---------------------------------------------------------------------- #

class TestDepoziteExcel:

    HEADERS = ["Nume depozit", "Latitudine", "Longitudine", "Tip depozit", "Status depozit", "Judet"]

    def test_valid_depozite(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            _create_xlsx(tmp, "Depozite_test.xlsx", self.HEADERS, [
                ["Dep A", "46.5", "24.3", "Depozit Temporar LR", "Activ", "Mures"],
                ["Dep B", "46.6", "24.4", "Depozit", "Activ", "Mures"]
            ])
            entries = reader.read_depozite_excel(tmp, "Depozite_test.xlsx")
            assert len(entries) == 2
            assert entries[0].nume_depozit == "Dep A"
            assert entries[1].tip_depozit == "Depozit"

    # ---------------------------------------------------------------------- #

    def test_empty_file_raises(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            wb = Workbook()
            wb.save(os.path.join(tmp, "empty.xlsx"))
            with pytest.raises(SRError, match = "gol"):
                reader.read_depozite_excel(tmp, "empty.xlsx")

    # ---------------------------------------------------------------------- #

    def test_missing_column_raises(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            _create_xlsx(tmp, "bad.xlsx", ["Wrong col"], [["val"]])
            with pytest.raises(SRError, match = "coloanele necesare"):
                reader.read_depozite_excel(tmp, "bad.xlsx")

# ---------------------------------------------------------------------- #

class TestIntrariExcel:

    HEADERS = ["Număr NIR", "Companie emitenta", "Cod aviz", "Dată NIR", "Grupe specii", "Volum (mc)", "Status"]

    def test_valid_intrari(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            _create_xlsx(tmp, "Intrari_test.xlsx", self.HEADERS, [
                ["1", "Firma A", "AP123", "2025-12-08 10:00:00.000000", "Rasinoase", "5.5", "Aprobat"]
            ])
            entries = reader.read_intrari_excel(tmp, "Intrari_test.xlsx")
            assert len(entries) == 1
            assert entries[0].numar_nir == 1
            assert entries[0].volum_mc == 5.5

# ---------------------------------------------------------------------- #

class TestAvizExcel:

    HEADERS = [
        "Cod Aviz", "Tip Transport", "Provenienta", "Mijloc transport",
        "Numar mijloc transport", "Categorie transport", "Status",
        "Volum total aviz(mc)", "Data emitere"
    ]

    def test_valid_aviz(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            _create_xlsx(tmp, "Aviz_test.xlsx", self.HEADERS, [
                ["AP456", "Intern", "Dep A", "Camion", "MS98LIR", "Cat A", "Valid", "10.0",
                 "2025-12-08 10:00:00.000000"]
            ])
            entries = reader.read_aviz_excel(tmp, "Aviz_test.xlsx")
            assert len(entries) == 1
            assert entries[0].cod_aviz == "AP456"

# ---------------------------------------------------------------------- #

class TestRowCount:

    def test_count_rows(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            _create_xlsx(tmp, "test.xlsx", ["A", "B"], [["1", "2"], ["3", "4"]])
            count = reader.read_excel_row_count(tmp, "test.xlsx")
            assert count == 2

    # ---------------------------------------------------------------------- #

    def test_count_returns_zero_for_missing(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            count = reader.read_excel_row_count(tmp, "nonexistent.xlsx")
            assert count == 0
    
    # ---------------------------------------------------------------------- #

    def test_count_returns_zero_for_empty(self, reader):
        with tempfile.TemporaryDirectory() as tmp:
            wb = Workbook()
            wb.save(os.path.join(tmp, "empty.xlsx"))
            count = reader.read_excel_row_count(tmp, "empty.xlsx")
            assert count == 0
