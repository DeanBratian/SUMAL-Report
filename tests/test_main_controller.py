import pytest
import os
import tempfile
from unittest.mock import MagicMock
from openpyxl import Workbook as XlWorkbook
from controllers.main_controller import MainController
from core.models import CSVParseResult, DepozitDataModel
from core.enums import NoticeType, DepositType, CSVParseStatus, FolderStatus, DepozitSource
from core.config import DEPOSIT_DATA_ENABLED_FIELDS_BY_TYPE
from core.sr_error import SRError
from tests.conftest import make_notice, make_deposit_data, make_depozite_excel_entry

@pytest.fixture
def ctrl():
    return MainController(MagicMock())

# ---------------------------------------------------------------------- #

class TestStateManagement:

    def test_set_folder_resets_state(self, ctrl):
        ctrl.is_parsed = True
        ctrl.notices = [MagicMock()]
        ctrl.set_folder("/some/path")
        assert ctrl.is_parsed is False
        assert ctrl.notices == []

    # ---------------------------------------------------------------------- #

    def test_set_same_folder_returns_false(self, ctrl):
        ctrl.set_folder("/path/a")
        assert ctrl.set_folder("/path/a") is False

    # ---------------------------------------------------------------------- #

    def test_set_different_folder_returns_true(self, ctrl):
        ctrl.set_folder("/path/a")
        assert ctrl.set_folder("/path/b") is True

# ---------------------------------------------------------------------- #

class TestCapabilityGates:

    def test_can_start_parsing_requires_all_conditions(self, ctrl):
        # Nothing set -> False
        assert not ctrl.can_start_parsing()

    # ---------------------------------------------------------------------- #

    def test_can_edit_deposit_data_requires_parsed(self, ctrl):
        assert ctrl.can_edit_deposit_data() is False

        ctrl.is_parsed = True
        ctrl.notices = [MagicMock()]
        ctrl.deposit_data = [MagicMock()]
        ctrl.csv_parse_results = [CSVParseResult("a.csv", CSVParseStatus.OK)]
        assert ctrl.can_edit_deposit_data() is True

    # ---------------------------------------------------------------------- #
    
    def test_can_edit_deposit_data_requires_parsed_errors(self, ctrl):
        assert ctrl.can_edit_deposit_data() is False

        ctrl.is_parsed = True
        ctrl.notices = [MagicMock()]
        ctrl.deposit_data = [MagicMock()]
        ctrl.csv_parse_results = [CSVParseResult("a.csv", CSVParseStatus.OK), CSVParseResult("b.csv", CSVParseStatus.ERROR)]
        assert ctrl.can_edit_deposit_data() is False

    # ---------------------------------------------------------------------- #

    def test_can_generate_reports_requires_complete_deposit_data(self, ctrl):
        assert ctrl.can_generate_reports() is False

    # ---------------------------------------------------------------------- #

    def test_can_generate_reports_with_valid_data(self, ctrl):
        ctrl.is_parsed = True
        ctrl.notices = [MagicMock()]
        ctrl.csv_parse_results = [CSVParseResult("a.csv", CSVParseStatus.OK)]

        dd = make_deposit_data(
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50,
            p_transport_mc = 30, p_iesire_lr_mc = 200,
            p_iesire_lf_mc = 80, p_expl_mc = 20
        )
        ctrl.deposit_data = [dd]
        assert ctrl.can_generate_reports() is True

    # ---------------------------------------------------------------------- #

    def test_can_generate_reports_with_partial_data(self, ctrl):
        ctrl.is_parsed = True
        ctrl.notices = [MagicMock()]
        ctrl.csv_parse_results = [CSVParseResult("a.csv", CSVParseStatus.OK)]

        dd = make_deposit_data(
            p_intrare_lr_mc = 0, p_intrare_lf_mc = 50,
            p_transport_mc = 30, p_iesire_lr_mc = 200,
            p_iesire_lf_mc = 80, p_expl_mc = 20
        )
        ctrl.deposit_data = [dd]
        assert ctrl.can_generate_reports() is False

# ---------------------------------------------------------------------- #

class TestInitializeDepositData:

    def test_filters_unreferenced_deposits(self, ctrl):
        ctrl.notices = [
            make_notice(provenienta = "Dep A"),
        ]
        ctrl.notices[0].type = NoticeType.INTRARE_DIN_PARTIDA_PROPRIE

        ctrl.depozite_entries = [
            make_depozite_excel_entry(nume_depozit = "Dep A", tip_depozit = "Depozit Temporar LR"),
            make_depozite_excel_entry(nume_depozit = "Dep B", tip_depozit = "Depozit")
        ]

        _, ignored = ctrl.initialize_deposit_data()

        assert len(ctrl.deposit_data) == 1
        assert ctrl.deposit_data[0].nume_depozit == "Dep A"
        assert "Dep B" in ignored

    # ---------------------------------------------------------------------- #

    def test_adds_external_deposits(self, ctrl):
        notice = make_notice(
            provenienta = "Firma Externă",
            emitent_nume = "Alt SRL", emitent_cui = "12345678",
            emitent_adresa = "Alta adresa"
        )
        notice.type = NoticeType.INTRARE_DIN_SURSA_EXTERNA
        ctrl.notices = [notice]
        ctrl.depozite_entries = []

        external, ignored = ctrl.initialize_deposit_data()

        assert "Firma Externă" in external
        assert len(ctrl.deposit_data) == 1
        assert ctrl.deposit_data[0].tip_depozit == DepositType.DEPOZIT_EXTERN

# ---------------------------------------------------------------------- #

class TestDepositDataValidation:

    def test_validate_empty_deposits_returns_false(self, ctrl):
        ctrl.deposit_data = []
        assert ctrl._validate_all_deposit_data_entries() is False
    
    # ---------------------------------------------------------------------- #

    def test_validate_incomplete_returns_false(self, ctrl):
        dd = make_deposit_data()  # all prices at 0
        ctrl.deposit_data = [dd]
        assert ctrl._validate_all_deposit_data_entries() is False

    # ---------------------------------------------------------------------- #

    def test_validate_complete_lr_returns_true(self, ctrl):
        dd = make_deposit_data(
            p_intrare_lr_mc = 100, p_intrare_lf_mc = 50,
            p_transport_mc = 30, p_iesire_lr_mc = 200,
            p_iesire_lf_mc = 80, p_expl_mc = 20
        )
        ctrl.deposit_data = [dd]
        assert ctrl._validate_all_deposit_data_entries() is True


# ---------------------------------------------------------------------- #

class TestTableData:

    def test_get_notice_returns_none_for_invalid_index(self, ctrl):
        assert ctrl.get_notice(0) is None
        assert ctrl.get_notice(-1) is None

    # ---------------------------------------------------------------------- #

    def test_get_notice_returns_correct(self, ctrl):
        n = make_notice(cod_unic="TEST")
        n.type = NoticeType.INTRARE_DIN_SURSA_EXTERNA
        ctrl.notices = [n]
        assert ctrl.get_notice(0).cod_unic == "TEST"

# ---------------------------------------------------------------------- #

class TestCsvCount:

    def test_csv_count_per_excels(self, ctrl):
        ctrl.intrari_xl_row_count = 10
        ctrl.aviz_xl_row_count = 5
        assert ctrl.get_csv_count_per_excels() == 15

    # ---------------------------------------------------------------------- #

    def test_is_correct_csv_count(self, ctrl):
        ctrl.intrari_xl_row_count = 3
        ctrl.aviz_xl_row_count = 2
        ctrl.csv_files_dict = {"all": ["a", "b", "c", "d", "e"], "unique": []}
        assert ctrl.is_correct_csv_count() is True

    # ---------------------------------------------------------------------- #

    def test_incorrect_csv_count(self, ctrl):
        ctrl.intrari_xl_row_count = 3
        ctrl.aviz_xl_row_count = 2
        ctrl.csv_files_dict = {"all": ["a", "b", "c"], "unique": []}
        assert ctrl.is_correct_csv_count() is False

# ---------------------------------------------------------------------- #

class TestResetState:

    def test_reset_clears_all_fields(self, ctrl):
        ctrl.is_parsed = True
        ctrl.notices = [MagicMock()]
        ctrl.intrari_entries = [MagicMock()]
        ctrl.aviz_entries = [MagicMock()]
        ctrl.depozite_entries = [MagicMock()]
        ctrl.csv_files_dict = {"unique": ["a.csv"], "all": ["a.csv"]}
        ctrl.xl_files_dict = {"intrari": "i.xlsx", "aviz": "a.xlsx", "depozite": "d.xlsx"}
        ctrl.intrari_xl_row_count = 5
        ctrl.aviz_xl_row_count = 3
        ctrl.depozite_xl_row_count = 2
        ctrl.csv_parse_results = [MagicMock()]
        ctrl.deposit_data = [MagicMock()]

        ctrl.reset_state()

        assert ctrl.is_parsed is False
        assert ctrl.notices == []
        assert ctrl.intrari_entries == []
        assert ctrl.aviz_entries == []
        assert ctrl.depozite_entries == []
        assert ctrl.csv_files_dict == {}
        assert ctrl.xl_files_dict == {}
        assert ctrl.intrari_xl_row_count == 0
        assert ctrl.aviz_xl_row_count == 0
        assert ctrl.depozite_xl_row_count == 0
        assert ctrl.csv_parse_results == []
        assert ctrl.deposit_data == []

# ---------------------------------------------------------------------- #

class TestErrorFiles:

    def test_no_errors_returns_empty_set(self, ctrl):
        ctrl.csv_parse_results = [
            CSVParseResult("a.csv", CSVParseStatus.OK),
            CSVParseResult("b.csv", CSVParseStatus.OK)
        ]
        assert ctrl.get_error_files() == set()

    # ---------------------------------------------------------------------- #

    def test_returns_error_filenames(self, ctrl):
        ctrl.csv_parse_results = [
            CSVParseResult("good.csv", CSVParseStatus.OK),
            CSVParseResult("bad.csv", CSVParseStatus.ERROR, "Parse failed"),
            CSVParseResult("bad2.csv", CSVParseStatus.ERROR, "Missing col")
        ]
        errors = ctrl.get_error_files()
        assert errors == {"bad.csv", "bad2.csv"}

# ---------------------------------------------------------------------- #

class TestDetectCsvFiles:

    def test_detects_csv_files(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            open(os.path.join(tmp, "notice1_08_12_2025_16_57_38.csv"), "w").close()
            open(os.path.join(tmp, "notice2_08_12_2025_16_57_38.csv"), "w").close()
            ctrl.folder_path = tmp
            ctrl.detect_csv_files()
            assert len(ctrl.csv_files_dict["all"]) == 2
            assert len(ctrl.csv_files_dict["unique"]) == 2

    # ---------------------------------------------------------------------- #

    def test_deduplicates_by_base_name(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            open(os.path.join(tmp, "AP123_08_12_2025_16_57_38.csv"), "w").close()
            open(os.path.join(tmp, "AP123_09_12_2025_10_00_00.csv"), "w").close()
            ctrl.folder_path = tmp
            ctrl.detect_csv_files()
            assert len(ctrl.csv_files_dict["all"]) == 2
            assert len(ctrl.csv_files_dict["unique"]) == 1

    # ---------------------------------------------------------------------- #

    def test_ignores_non_csv_files(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            open(os.path.join(tmp, "data.xlsx"), "w").close()
            open(os.path.join(tmp, "readme.txt"), "w").close()
            ctrl.folder_path = tmp
            ctrl.detect_csv_files()
            assert len(ctrl.csv_files_dict["all"]) == 0

    # ---------------------------------------------------------------------- #

    def test_raises_on_invalid_folder(self, ctrl):
        ctrl.folder_path = "/nonexistent/path/here"
        with pytest.raises(SRError):
            ctrl.detect_csv_files()


# ---------------------------------------------------------------------- #

class TestDetectExcelFiles:

    def test_detects_all_three_excel_files(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            from openpyxl import Workbook
            for name in ["Intrari_test.xlsx", "Aviz_test.xlsx", "Depozite_test.xlsx"]:
                wb = Workbook()
                ws = wb.active
                ws.append(["col1"])
                ws.append(["val1"])
                wb.save(os.path.join(tmp, name))

            ctrl.folder_path = tmp
            result = ctrl.detect_excel_files()
            assert result is True
            assert "intrari" in ctrl.xl_files_dict
            assert "aviz" in ctrl.xl_files_dict
            assert "depozite" in ctrl.xl_files_dict

            assert ctrl.xl_files_dict["intrari"] == "Intrari_test.xlsx"
            assert ctrl.xl_files_dict["aviz"] == "Aviz_test.xlsx"
            assert ctrl.xl_files_dict["depozite"] == "Depozite_test.xlsx"

    # ---------------------------------------------------------------------- #

    def test_returns_false_when_missing_excel(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            ctrl.folder_path = tmp
            result = ctrl.detect_excel_files()
            assert result is False

    # ---------------------------------------------------------------------- #

    def test_raises_on_invalid_folder(self, ctrl):
        ctrl.folder_path = "/nonexistent/path/here"
        with pytest.raises(SRError):
            ctrl.detect_excel_files()

    # ---------------------------------------------------------------------- #

    def test_raises_on_duplicate_excel_files(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            for name in ["Intrari_test.xlsx", "Intrari_test (1).xlsx",
                         "Aviz_test.xlsx", "Depozite_test.xlsx"]:
                open(os.path.join(tmp, name), "w").close()
            ctrl.folder_path = tmp
            with pytest.raises(SRError) as exc_info:
                ctrl.detect_excel_files()
            assert "Intrari_test.xlsx" in exc_info.value.message
            assert "Intrari_test (1).xlsx" in exc_info.value.message
            assert ctrl.xl_files_dict == {}

# ---------------------------------------------------------------------- #

class TestCanStartParsingFull:

    def test_all_conditions_met(self, ctrl):
        ctrl.folder_path = "/some/path"
        ctrl.depozite_xl_row_count = 2
        ctrl.csv_files_dict = {"unique": ["a.csv"], "all": ["a.csv", "b.csv", "c.csv"]}
        ctrl.xl_files_dict = {"intrari": "i.xlsx", "aviz": "a.xlsx", "depozite": "d.xlsx"}
        ctrl.intrari_xl_row_count = 2
        ctrl.aviz_xl_row_count = 1
        assert ctrl.can_start_parsing()

    # ---------------------------------------------------------------------- #

    def test_missing_folder(self, ctrl):
        ctrl.folder_path = None
        ctrl.depozite_xl_row_count = 2
        ctrl.csv_files_dict = {"unique": ["a.csv"], "all": ["a.csv"]}
        ctrl.xl_files_dict = {"intrari": "i.xlsx", "aviz": "a.xlsx", "depozite": "d.xlsx"}
        ctrl.intrari_xl_row_count = 1
        ctrl.aviz_xl_row_count = 0
        assert not ctrl.can_start_parsing()

    # ---------------------------------------------------------------------- #

    def test_missing_depozite_rows(self, ctrl):
        # Everything else valid — only the empty Depozite_ file should block parsing,
        # so this isolates the depozite_xl_row_count gate.
        ctrl.folder_path = "/some/path"
        ctrl.depozite_xl_row_count = 0
        ctrl.csv_files_dict = {"unique": ["a.csv"], "all": ["a.csv"]}
        ctrl.xl_files_dict = {"intrari": "i.xlsx", "aviz": "a.xlsx", "depozite": "d.xlsx"}
        ctrl.intrari_xl_row_count = 1
        ctrl.aviz_xl_row_count = 0
        assert not ctrl.can_start_parsing()

    # ---------------------------------------------------------------------- #

    def test_csv_count_mismatch(self, ctrl):
        ctrl.folder_path = "/some/path"
        ctrl.depozite_xl_row_count = 2
        ctrl.csv_files_dict = {"unique": ["a.csv"], "all": ["a.csv"]}
        ctrl.xl_files_dict = {"intrari": "i.xlsx", "aviz": "a.xlsx", "depozite": "d.xlsx"}
        ctrl.intrari_xl_row_count = 5
        ctrl.aviz_xl_row_count = 5
        assert not ctrl.can_start_parsing()

    # ---------------------------------------------------------------------- #

    def test_already_parsed(self, ctrl):
        ctrl.folder_path = "/some/path"
        ctrl.depozite_xl_row_count = 2
        ctrl.csv_files_dict = {"unique": ["a.csv"], "all": ["a.csv"]}
        ctrl.xl_files_dict = {"intrari": "i.xlsx", "aviz": "a.xlsx", "depozite": "d.xlsx"}
        ctrl.intrari_xl_row_count = 1
        ctrl.aviz_xl_row_count = 0
        ctrl.is_parsed = True
        assert not ctrl.can_start_parsing()

# ---------------------------------------------------------------------- #

class TestGenerateReport:

    def test_raises_on_missing_deposit_for_notice(self, ctrl):
        notice = make_notice(provenienta="Unknown Deposit")
        notice.type = NoticeType.INTRARE_DIN_SURSA_EXTERNA
        ctrl.notices = [notice]
        ctrl.deposit_data = [make_deposit_data(name = "Different Deposit")]

        with pytest.raises(SRError, match = "Nu s-au găsit date"):
            ctrl.generate_report("/tmp/output.xlsx", [])

# ---------------------------------------------------------------------- #

class TestInitializeDepositDataEdgeCases:

    def test_unknown_deposit_type_raises(self, ctrl):
        notice = make_notice(provenienta = "Dep X")
        notice.type = NoticeType.INTRARE_DIN_PARTIDA_PROPRIE
        ctrl.notices = [notice]
        ctrl.depozite_entries = [
            make_depozite_excel_entry(nume_depozit = "Dep X", tip_depozit = "Tip Inventat")
        ]
        with pytest.raises(SRError, match = "nu a putut fi determinat"):
            ctrl.initialize_deposit_data()

    # ---------------------------------------------------------------------- #

    def test_clears_previous_deposit_data(self, ctrl):
        ctrl.deposit_data = [make_deposit_data(name = "Old")]
        ctrl.notices = [make_notice(provenienta = "New")]
        ctrl.notices[0].type = NoticeType.INTRARE_DIN_PARTIDA_PROPRIE
        ctrl.depozite_entries = [
            make_depozite_excel_entry(nume_depozit = "New", tip_depozit = "Depozit Temporar LR")
        ]
        ctrl.initialize_deposit_data()
        assert len(ctrl.deposit_data) == 1
        assert ctrl.deposit_data[0].nume_depozit == "New"

    # ---------------------------------------------------------------------- #

    def test_mixed_excel_and_external_deposits(self, ctrl):
        n1 = make_notice(provenienta = "Dep Excel")
        n1.type = NoticeType.INTRARE_DIN_PARTIDA_PROPRIE
        n2 = make_notice(
            provenienta = "Firma Externa",
            emitent_nume = "Alt SRL", emitent_cui = "12345678",
            emitent_adresa = "Alta adresa",
        )
        n2.type = NoticeType.INTRARE_DIN_SURSA_EXTERNA
        ctrl.notices = [n1, n2]
        ctrl.depozite_entries = [
            make_depozite_excel_entry(nume_depozit = "Dep Excel", tip_depozit = "Depozit Temporar LR")
        ]

        external, ignored = ctrl.initialize_deposit_data()

        assert len(ctrl.deposit_data) == 2
        names = {d.nume_depozit for d in ctrl.deposit_data}
        assert "Dep Excel" in names
        assert "Firma Externa" in names
        assert "Firma Externa" in external

# ---------------------------------------------------------------------- #

class TestTableDataStructure:

    def test_table_data_has_expected_keys(self, ctrl):
        n = make_notice(cod_unic = "ABC123")
        n.type = NoticeType.IESIRE_DIN_DEPOZIT_PRINCIPAL
        ctrl.notices = [n]
        table_data = ctrl.get_main_window_table_data()

        row = table_data[0]
        assert set(row.keys()) == {"cod_unic", "data", "emitent", "destinatar", "volum", "tip"}
        assert row["cod_unic"] == "ABC123"
        assert "m³" in row["volum"]

    # ---------------------------------------------------------------------- #

    def test_empty_notices_returns_empty_list(self, ctrl):
        ctrl.notices = []
        assert ctrl.get_main_window_table_data() == []

# ---------------------------------------------------------------------- #

class TestValidateFolderFiles:

    def test_no_files_status(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            ctrl.folder_path = tmp
            status, data = ctrl.validate_folder_files()
            assert status == FolderStatus.NO_FILES

    # ---------------------------------------------------------------------- #

    def test_missing_xl_status(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            open(os.path.join(tmp, "notice_08_12_2025_16_57_38.csv"), "w").close()
            ctrl.folder_path = tmp
            status, data = ctrl.validate_folder_files()
            assert status == FolderStatus.MISSING_XL

    # ---------------------------------------------------------------------- #

    def test_missing_csv_status(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            for name in ["Intrari_test.xlsx", "Aviz_test.xlsx", "Depozite_test.xlsx"]:
                wb = XlWorkbook()
                ws = wb.active
                ws.append(["col1"])
                ws.append(["val1"])
                wb.save(os.path.join(tmp, name))
            # intrari(1) + aviz(1) = expects 2 CSVs, only 1 provided
            open(os.path.join(tmp, "notice_08_12_2025_16_57_38.csv"), "w").close()
            ctrl.folder_path = tmp
            status, data = ctrl.validate_folder_files()
            assert status == FolderStatus.MISSING_CSV

    # ---------------------------------------------------------------------- #

    def test_ready_status(self, ctrl):
        with tempfile.TemporaryDirectory() as tmp:
            for name in ["Intrari_test.xlsx", "Aviz_test.xlsx", "Depozite_test.xlsx"]:
                wb = XlWorkbook()
                ws = wb.active
                ws.append(["col1"])
                ws.append(["val1"])
                wb.save(os.path.join(tmp, name))
            # intrari(1) + aviz(1) = 2 CSVs expected
            open(os.path.join(tmp, "n1_08_12_2025_16_57_38.csv"), "w").close()
            open(os.path.join(tmp, "n2_08_12_2025_16_57_38.csv"), "w").close()
            ctrl.folder_path = tmp
            status, data = ctrl.validate_folder_files()
            assert status == FolderStatus.READY

# ---------------------------------------------------------------------- #

class TestValidateDepositDataByType:

    def test_depozit_principal_complete(self, ctrl):
        dd = DepozitDataModel("Main", "Depozit", DepozitSource.EXCEL_SUMAL)
        dd.price_data.p_iesire_lr_mc = 200
        dd.price_data.p_iesire_lf_mc = 80
        dd.price_data.p_iesire_ch_mc = 60
        dd.price_data.p_transport_mc = 40
        ctrl.deposit_data = [dd]
        assert ctrl._validate_all_deposit_data_entries() is True

    # ---------------------------------------------------------------------- #

    def test_depozit_principal_missing_field(self, ctrl):
        dd = DepozitDataModel("Main", "Depozit", DepozitSource.EXCEL_SUMAL)
        dd.price_data.p_iesire_lr_mc = 200
        dd.price_data.p_iesire_lf_mc = 80
        # missing p_iesire_ch_mc and p_transport_mc
        ctrl.deposit_data = [dd]
        assert ctrl._validate_all_deposit_data_entries() is False

    # ---------------------------------------------------------------------- #

    def test_depozit_extern_complete(self, ctrl):
        dd = DepozitDataModel("Ext", "Depozit extern", DepozitSource.AVIZE)
        dd.price_data.p_intrare_lr_mc = 100
        dd.price_data.p_intrare_lf_mc = 50
        dd.price_data.p_intrare_ch_mc = 80
        dd.price_data.p_transport_mc = 30
        ctrl.deposit_data = [dd]
        assert ctrl._validate_all_deposit_data_entries() is True

    # ---------------------------------------------------------------------- #

    def test_one_incomplete_fails_all(self, ctrl):
        dd_ok = DepozitDataModel("LR", "Depozit Temporar LR", DepozitSource.EXCEL_SUMAL)
        for f in DEPOSIT_DATA_ENABLED_FIELDS_BY_TYPE[DepositType.DEPOZIT_TEMPORAR_LR]:
            setattr(dd_ok.price_data, f, 10)

        dd_bad = DepozitDataModel("Ext", "Depozit extern", DepozitSource.AVIZE)
        # all prices at 0
        ctrl.deposit_data = [dd_ok, dd_bad]
        assert ctrl._validate_all_deposit_data_entries() is False
